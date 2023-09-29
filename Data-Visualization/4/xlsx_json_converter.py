import datetime
import json
import openpyxl

sheet_input = openpyxl.load_workbook("CityData.xlsx")["CityData"]

data = []
for row in range(1, sheet_input.max_row + 1):
	provinceName = (sheet_input.cell(row, 2)).value
	cityName = (sheet_input.cell(row, 3)).value
	city_confirmedCount = (sheet_input.cell(row, 4)).value
	updateTime = str((sheet_input.cell(row, 8)).value)[0:10]
	data.append([provinceName, cityName, city_confirmedCount, updateTime])

print(data[0], data[1], data[-1])

with open("CityData.js", "w") as json_output:
	json.dump(data, json_output, indent = "  ")